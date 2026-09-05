"""Lee únicamente filas pobladas de la planilla maestra."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .normalize import (
    canon_company,
    canon_partner,
    money,
    norm_ncf,
    norm_vat,
    ncf_prefix,
    ncf_seq,
)


def _cell_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if text.upper() in {"N/A", "NA", "NO APLICA"}:
        return text
    return text


def _row_has_data(values: tuple[Any, ...], required_cols: int = 3) -> bool:
    filled = [v for v in values[:required_cols] if v not in (None, "", " ")]
    return len(filled) >= 1 and any(v not in (None, "", " ") for v in values)


def parse_workbook(path: str | Path) -> dict:
    wb = load_workbook(path, data_only=True)
    out = {
        "source": str(path),
        "sheets": {},
        "users": [],
        "ncf_sequences": [],
        "cxc": [],
        "cxp": [],
        "banks_pending": [],
        "customers_master": [],
        "vendors_master": [],
        "fixed_assets": [],
        "cutover": {},
    }
    out["sheets"] = {name: {"max_row": wb[name].max_row} for name in wb.sheetnames}

    users_ws = wb["01_Usuarios"]
    for row in users_ws.iter_rows(min_row=2, values_only=True):
        if not any(c not in (None, "") for c in row[:8]):
            continue
        out["users"].append(
            {
                "name": row[0],
                "login": row[1],
                "phone": row[2],
                "role": row[3],
                "company": row[4],
            }
        )

    ncf_ws = wb["02_Secuencias_NCF"]
    for row in ncf_ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        declared = str(row[1]).strip().upper()
        last = norm_ncf(row[5]) if row[5] else ""
        nxt = norm_ncf(row[6]) if row[6] else ""
        rng_from = norm_ncf(row[3]) if row[3] else ""
        rng_to = norm_ncf(row[4]) if row[4] else ""
        conflicts = []
        if last and ncf_prefix(last) != declared:
            conflicts.append("LAST_PREFIX_NE_TYPE")
        if nxt and ncf_prefix(nxt) != declared:
            conflicts.append("NEXT_PREFIX_NE_TYPE")
        if rng_from and ncf_prefix(rng_from) != declared:
            conflicts.append("RANGE_FROM_PREFIX_NE_TYPE")
        if rng_to and ncf_prefix(rng_to) != declared:
            conflicts.append("RANGE_TO_PREFIX_NE_TYPE")
        last_seq = ncf_seq(last)
        next_seq = ncf_seq(nxt)
        from_seq = ncf_seq(rng_from)
        to_seq = ncf_seq(rng_to)
        if last_seq is not None and next_seq is not None and next_seq <= last_seq:
            conflicts.append("NEXT_NOT_GREATER_THAN_LAST")
        if next_seq is not None and from_seq is not None and next_seq < from_seq:
            conflicts.append("NEXT_BELOW_RANGE")
        if next_seq is not None and to_seq is not None and next_seq > to_seq:
            conflicts.append("NEXT_ABOVE_RANGE")
        if last_seq is not None and from_seq is not None and last_seq < from_seq:
            conflicts.append("LAST_BELOW_DECLARED_RANGE")
        out["ncf_sequences"].append(
            {
                "company": canon_company(str(row[0])),
                "excel_company": row[0],
                "declared_type": declared,
                "range_from": rng_from,
                "range_to": rng_to,
                "last_used": last,
                "next": nxt,
                "expiration": _cell_date(row[7]),
                "authorization": str(row[8]) if row[8] not in (None, "") else "",
                "conflicts": conflicts,
                "status": "NCF_SEQUENCE_CONFLICT" if conflicts else "CONSISTENT",
            }
        )

    cxc_ws = wb["03_Cuentas_por_Cobrar"]
    for idx, row in enumerate(cxc_ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not (row[3] or row[4]):
            continue
        original = money(row[8])
        paid = money(row[9])
        residual = money(row[10])
        expected = original - paid
        balance_ok = expected == residual
        out["cxc"].append(
            {
                "excel_row": idx,
                "company": canon_company(str(row[0])),
                "excel_company": row[0],
                "customer": canon_partner(str(row[1] or "")),
                "excel_customer": row[1],
                "vat": norm_vat(row[2]),
                "vat_display": str(row[2] or ""),
                "invoice_number": str(row[3] or ""),
                "ncf": norm_ncf(row[4] or row[3]),
                "invoice_date": _cell_date(row[5]),
                "due_date": _cell_date(row[6]),
                "currency": (row[7] or "DOP"),
                "amount_original": str(original),
                "amount_paid": str(paid),
                "amount_residual": str(residual),
                "balance_ok": balance_ok,
                "pdf_name": row[11],
                "notes": row[12],
                "excel_state": row[13],
            }
        )

    cxp_ws = wb["04_Cuentas_por_Pagar"]
    for idx, row in enumerate(cxp_ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not (row[3] or row[4]):
            continue
        out["cxp"].append({"excel_row": idx, "company": row[0], "ncf": row[4]})

    cust_ws = wb["06_Clientes"]
    group_names = {
        canon_company(c)
        for c in (
            "INVERSIONES DORALEX,S.RL.",
            "COMERCIALIZADORA DE ALIMENTOS PIÑARIA, S.R.L.",
            "DOMINION BUSINESS,S.R.L.",
            "INVERSIONES EL MAYUMA, S.R.L.",
            "REMPART GROUP S.R.L.",
            "BLUE ELITE, S.R.L.",
        )
    }
    for row in cust_ws.iter_rows(min_row=2, values_only=True):
        if not row[0] and not row[1]:
            continue
        name = str(row[1] or "")
        vat = norm_vat(row[2])
        if not vat and canon_company(name) in group_names:
            continue
        if not vat and not name:
            continue
        if not vat and canon_company(str(row[0] or "")) == canon_company(name):
            continue
        out["customers_master"].append({"company": row[0], "name": name, "vat": vat})

    vend_ws = wb["07_Proveedores"]
    for row in vend_ws.iter_rows(min_row=2, values_only=True):
        if not any(c not in (None, "") for c in row[:3]):
            continue
        out["vendors_master"].append({"company": row[0], "name": row[1], "vat": row[2]})

    ast_ws = wb["08_Activos_Fijos"]
    for row in ast_ws.iter_rows(min_row=2, values_only=True):
        if not any(c not in (None, "") for c in row[:4]):
            continue
        out["fixed_assets"].append({"company": row[0], "description": row[2]})

    cut_ws = wb["09_Corte_y_Confirmaciones"]
    rows = list(cut_ws.iter_rows(values_only=True))
    for row in rows[2:]:
        if not row[0]:
            continue
        out["cutover"][str(row[0])] = {
            "answer": _cell_date(row[1]) if not isinstance(row[1], str) else row[1],
            "note": row[3],
        }

    return out


def ar_totals(cxc: list[dict]) -> dict[str, Decimal]:
    totals: dict[str, Decimal] = {}
    for row in cxc:
        company = row["company"]
        totals[company] = totals.get(company, Decimal("0")) + money(
            row["amount_residual"]
        )
    return totals
